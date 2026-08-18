"""数据读写 -- 支持 CSV / Excel / Parquet，自动解析路径。

扩展自 2024-A 项目中 openpyxl 写出模式，新增通用矩阵读写与格式化。
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from numpy import ndarray

DATA_DIR = Path(__file__).resolve().parents[1] / "data"

_EXCEL_ENGINE = "openpyxl"


# ---------------------------------------------------------------------------
# DataFrame I/O（保留原有接口）
# ---------------------------------------------------------------------------


def load_raw(filename: str) -> pd.DataFrame:
    """从 data/raw/ 加载数据集。"""
    path = DATA_DIR / "raw" / filename
    if path.suffix == ".csv":
        return pd.read_csv(path)
    if path.suffix in (".xlsx", ".xls"):
        return pd.read_excel(path, engine=_EXCEL_ENGINE)
    if path.suffix == ".parquet":
        return pd.read_parquet(path)
    raise ValueError(f"不支持的格式：{path.suffix}")


def save_processed(df: pd.DataFrame, filename: str) -> None:
    """保存处理后的 DataFrame 到 data/processed/。"""
    path = DATA_DIR / "processed" / filename
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.suffix == ".csv":
        df.to_csv(path, index=False)
    elif path.suffix == ".parquet":
        df.to_parquet(path, index=False)
    elif path.suffix in (".xlsx", ".xls"):
        df.to_excel(path, index=False, engine=_EXCEL_ENGINE)
    else:
        raise ValueError(f"不支持的格式：{path.suffix}")
    print(f"已保存：{path}")


# ---------------------------------------------------------------------------
# 矩阵读写（Excel ↔ ndarray）
# ---------------------------------------------------------------------------


def read_matrix_excel(path: str, sheet_name: str | int = 0) -> ndarray:
    """从 Excel 读取矩阵，自动跳过首行首列中的非数值标签。

    Args:
        path: Excel 文件路径。
        sheet_name: sheet 名称或索引。

    Returns:
        (n, m) 的 float 数组。
    """
    df = pd.read_excel(path, sheet_name=sheet_name, header=None, engine=_EXCEL_ENGINE)

    # 若首列全是字符串标签，丢弃首列
    try:
        df.iloc[:, 0].astype(float)
    except (ValueError, TypeError):
        df = df.iloc[:, 1:]

    # 若首行包含非数值，丢弃首行
    try:
        df.iloc[0].astype(float)
    except (ValueError, TypeError):
        df = df.iloc[1:]

    return df.to_numpy(dtype=float)


def write_matrix_excel(
    matrix: ndarray,
    path: str,
    *,
    sheet_name: str = "Sheet1",
    row_labels: list[str] | None = None,
    col_labels: list[str] | None = None,
    title: str | None = None,
) -> None:
    """将矩阵写入 Excel。

    Args:
        matrix: 数值矩阵 ``(n, m)``。
        path: 输出 .xlsx 文件路径。
        sheet_name: sheet 名称。
        row_labels: 行标签（放首列）。
        col_labels: 列标签（放首行）。
        title: 可选的表头标题（放在 A1 单元格上方一行）。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    n, m = matrix.shape
    has_title = title is not None
    has_col = col_labels is not None
    has_row = row_labels is not None
    offset_row = 0
    offset_col = 0

    # 标题行
    if has_title:
        ws.cell(1, 1, title)
        ws.cell(1, 1).font = Font(bold=True, size=12)
        offset_row = 1

    # 列标签
    if has_col:
        if has_row:
            ws.cell(offset_row + 1, 1)  # 左上角空单元格
            offset_col = 1
        for j, label in enumerate(col_labels):
            cell = ws.cell(offset_row + 1, offset_col + j + 1, label)
            cell.font = Font(bold=True)

    # 行标签 + 数据
    data_start_row = offset_row + 1 if (has_title or has_col) else 0
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    for i in range(n):
        row = data_start_row + 1 + i
        if has_row:
            cell = ws.cell(row, 1, row_labels[i])
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for j in range(m):
            val = float(matrix[i, j])
            cell = ws.cell(row, offset_col + j + 1, val)
            cell.number_format = "0.0000"
            cell.alignment = Alignment(horizontal="center")

    # 列宽自适应
    _auto_width(ws)

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"矩阵已写入：{path}  [{n}x{m}]")


# ---------------------------------------------------------------------------
# 多 Sheet Excel 写入（参考 2024-A 模式）
# ---------------------------------------------------------------------------


def write_sheets(
    sheets: dict[str, ndarray | pd.DataFrame],
    path: str,
    *,
    float_fmt: str = "%.4f",
) -> None:
    """一次写出含多个 sheet 的 Excel 文件。

    Args:
        sheets: ``{sheet名: 矩阵或DataFrame}`` 字典。
        path: 输出 .xlsx 路径。
        float_fmt: 数值格式。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment

    wb = Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    for name, data in sheets.items():
        ws = wb.create_sheet(name)
        if isinstance(data, pd.DataFrame):
            _write_df(ws, data, float_fmt)
        else:
            _write_ndarray(ws, data, float_fmt)
        _auto_width(ws)

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"已写出 {len(sheets)} 个 sheet -> {path}")


# ---------------------------------------------------------------------------
# 辅助
# ---------------------------------------------------------------------------


def _write_ndarray(ws, arr: ndarray, float_fmt: str) -> None:
    from openpyxl.styles import Alignment

    for i in range(arr.shape[0]):
        for j in range(arr.shape[1]):
            cell = ws.cell(i + 1, j + 1, float(arr[i, j]))
            cell.number_format = "0.0000"
            cell.alignment = Alignment(horizontal="center")


def _write_df(ws, df: pd.DataFrame, float_fmt: str) -> None:
    from openpyxl.styles import Alignment, Font

    # header
    for j, col in enumerate(df.columns):
        cell = ws.cell(1, j + 1, str(col))
        cell.font = Font(bold=True)
    # data
    for i in range(len(df)):
        for j in range(len(df.columns)):
            val = df.iloc[i, j]
            cell = ws.cell(i + 2, j + 1, val if not isinstance(val, float) else float(val))
            if isinstance(val, float):
                cell.number_format = "0.0000"
            cell.alignment = Alignment(horizontal="center")


def _auto_width(ws) -> None:
    """简单列宽自适应。"""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)
